import chromedriver_binary
import datetime
import requests
import re
import json
import time
import openpyxl
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import os
from flask import Flask, request, abort
from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    InvalidSignatureError
)
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage
)

app = Flask(__name__)

class UserContext:
    waiting_for_reset_number = False
user_context = {}

class wholeapp:
    f = 0
w = wholeapp()

# LINE Developersのチャネルシークレットとチャネルアクセストークンを設定
LINE_CHANNEL_ACCESS_TOKEN = os.getenv('LINE_CHANNEL_ACCESS_TOKEN')
LINE_CHANNEL_SECRET = os.getenv('LINE_CHANNEL_SECRET')
access_token = os.getenv('LINE_Notify_token')
line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)
LINEurl = "https://notify-api.line.me/api/notify"

excel_path = '/Users/aki_u/documents/Python/buslocate.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb['Sheet1']
wb.save(excel_path)
wb.close

class dettime:
    dettime1361 = "2024/02/25-03:00"
    dettime1362 = "2024/02/25-03:00"
d = dettime()

courseIDListAyase = ["041","231","177","217","170","212","131","092","102","140","202","054","064","166","105","148","112","180","110","085","248","247","046","121","019","039","135","199","070","152","002","089","107","410","409","049","136","211","167","078","060","141","057","082","127","214"]
courseIDListAsahi = ["130","207","058","100","129","149","171","023","153","042","138","145","222","108","010","133","517","516","513","514","000","386","384","391","224","372","370","392","000","492","371","405","390","368","377","376","387","383","366","380","491","489","000","354","337","320","344","040","114","086","164","219","000","001","532","113","174","034","096","009","144","003","074","190","155","073","098","175","147","168","188","000","388","375","097","436","438","427","242","238","430","429","426","434","432","000","523","519","529","518","146","183","526","522","530","515","528","520","524","512","545","546","000","582","583","584","585"]



def extract_js_var(soup, js_var):
    script = soup.find('script', string=re.compile(js_var, re.DOTALL))    #scriptタグを検索
    if script is not None:
        regex = '(?:var )' + js_var + '(.*?)' + '(?: = new navitime.geo.overlay.Pin(.*?)({[\s\S]*})(.*?);)' #busPinで検索
        json_str = re.search(regex, script.string).group(1) #最上位(始発に近い)のピンを選択
        if "next" in json_str:
            json_str = json_str.replace('_next','')
        return json.loads(json_str)

def log():
    FMT = "%Y/%m/%d-%H:%M"
    now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9))).strftime(FMT)
    FMTtime = "%H:%M"
    nowtime = datetime.datetime.now().strftime(FMTtime)
    td = datetime.timedelta(hours=1, minutes=30)
    token_dic = {'Authorization': 'Bearer ' + access_token}
    MaxRow = ws.max_row +1
    ws.cell(MaxRow,1,value = nowtime)
    ws.cell(MaxRow,22,value = nowtime)
    ws.cell(MaxRow,44,value = nowtime)
    ws.cell(MaxRow,54,value = nowtime)
    ws.cell(MaxRow,73,value = nowtime)
    ws.cell(MaxRow,87,value = nowtime)
    ws.cell(MaxRow,104,value = nowtime)





    for a in range(0,len(courseIDListAyase)):
        url_locate = 'https://transfer.navitime.biz/sotetsu/smart/location/BusLocationMap?courseId=0003400'+ courseIDListAyase[a]
        res = requests.get(url_locate)
        
        if __name__ == "__main__":
            html = res.text
            soup = bs(html, 'html.parser')
            td1361 = datetime.datetime.strptime(now,FMT) - datetime.datetime.strptime(d.dettime1361,FMT)
            td1362 = datetime.datetime.strptime(now,FMT) - datetime.datetime.strptime(d.dettime1362,FMT)
            CarNum = extract_js_var(soup, 'busPin')
            ws.cell(MaxRow,a,value = CarNum)
            url_navi_loc = 'https://transfer.navitime.biz/sotetsu-style-contents/bus-location/stops?courseId=0003400' + courseIDListAyase[a] + '&vehicleId=' + str(CarNum)
            if CarNum == 1361:
                if td1361 > td:
                    message1361 = '1361が動いています。' + str(a+1) +'列目です\n' + url_navi_loc
                    payload1361 = {'message': message1361}
                    requests.post(LINEurl, headers=token_dic, data=payload1361,)
                d.dettime1361 = now
            if CarNum == 1362:
                if td1362 > td:
                    message1362 = '1362が動いています。' + str(a+1) +'列目です\n' + url_navi_loc
                    payload1362 = {'message': message1362}
                    requests.post(LINEurl, headers=token_dic, data=payload1362,)
                d.dettime1362 = now
            time.sleep(0.3)
    wb.save(excel_path)



@app.route('/notify', methods=['HEAD','GET'])
def notify():
    log()
    return'OK'
    

@app.route("/callback", methods=['POST'])
def callback():
    # X-Line-Signatureヘッダーの取得
    signature = request.headers['X-Line-Signature']

    # リクエストの取得
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)

    # 署名の検証
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return 'OK'

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    # 受信したメッセージを取得
    received_text = event.message.text
    if received_text == "Reset":
        user_id = event.source.user_id
        user_context[user_id] = UserContext()
        user_context[user_id].waiting_for_reset_number = True
        line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="リセットする車番を入力してください：")
        )

    elif received_text.isdigit() and user_context.get(event.source.user_id) and user_context[event.source.user_id].waiting_for_reset_number:
        user_context[event.source.user_id].waiting_for_reset_number = False
        CarNum = str(received_text)
        if CarNum == "1361":
            d.dettime1361 = "03:00"
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="1361をリセットしました")
            )

        if CarNum == "1362":
            d.dettime1362 = "03:00"
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text="1362をリセットしました")
            )

    elif received_text == "dettime1361":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=d.dettime1361)
        )

    elif received_text == "dettime1362":
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=d.dettime1362)
        )
    
    else:
        user_id = event.source.user_id
        user_context[user_id] = UserContext()
        user_context[user_id].waiting_for_reset_number = False
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="動作停止中：")
        )

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)
